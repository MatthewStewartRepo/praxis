from mimetypes import init
import openpyxl
import numpy
import pathlib
import pyautogui
import statistics
import datetime




def importFile():
    start_time = datetime.datetime.now().replace(microsecond=0)
    print("Script initiated...")
    #find file in certain location 
    current = pathlib.Path(r"C:\Users\Matts\Documents\GWU\Praxis\DataFiles\Reaper-Crypt-Analysis_filtered.csv")
    workbook = openpyxl.load_workbook(r'C:\\Users\\Matts\\Documents\\GWU\\Praxis\\DataFiles\\betterPraxisData.xlsx')
    counter = len(workbook.sheetnames)
    print("There are %d sheets in the file..." %counter)
    sheets = workbook.sheetnames
    print("The worksheet names are:", sheets)
    print("----------")
    worksheet = workbook.active
    for row in worksheet.iter_cols():
        if row[0].value is None:
            break
        print(row[0].value)
    print("----------")
    worksheet = workbook["User Categories"]
    print("Data for worksheet: User Categories")
    for row in worksheet.iter_cols():
        if row[0].value is None:
            break
        print(row[0].value)
    print("----------")
    print("Data for worksheet: Wallet Stats")
    worksheet = workbook["Wallet Stats"]
    for row in worksheet.iter_cols():
        if row[0].value is None:
            break
        print(row[0].value)
    print("----------")
    current_time = datetime.datetime.now().replace(microsecond=0) - start_time
    print("Time ran: ", current_time)



            

importFile()
        
        

